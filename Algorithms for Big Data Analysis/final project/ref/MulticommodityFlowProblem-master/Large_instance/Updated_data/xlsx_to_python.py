
# --- TRANSFORM XLSX DATA TO PYTHON ARRAY ---

import openpyxl
import numpy

# Get Arcs and Nodes
excelfile = openpyxl.load_workbook('/Users/nicolas/Documents/EPFL/4.Printemps2018/Project/Column_generation/Large_instance/New_data/SemesterProject_MFP_MediumInstance_Updated.xlsx')  # open medium data excel file
sheet1 = excelfile.get_sheet_by_name("Nodes_Final") # Open each sheet
sheet2 = excelfile.get_sheet_by_name("Arcs_Final")

# Get Commodities
excelfile2 = openpyxl.load_workbook('/Users/nicolas/Documents/EPFL/4.Printemps2018/Project/Column_generation/Large_instance/New_data/SemesterProject_MFP_LargeInstance.xlsx')  # open medium data excel fil
sheet3 = excelfile2.get_sheet_by_name("OD_Matrix_Final")

# Create new .txt file to write the transformed data
newtext = open("/Users/nicolas/Documents/EPFL/4.Printemps2018/Project/Column_generation/Large_instance/New_data/data.txt","w")

# --- NODES ---

newtext.write('node = [ \n\t')

for row_index in range(2, (sheet1.max_row+1)):
    newtext.write('[\'' + str(sheet1.cell(row=row_index, column=1).value)  + '\', ')   # Get station name
    newtext.write(str(sheet1.cell(row=row_index, column=1).value[1:].lstrip('0')) + ', ')   # Get station id
    if(str(sheet1.cell(row=row_index, column=2).value) == 'None'):  #Transform station type into integers
        newtext.write('1')
    if(str(sheet1.cell(row=row_index, column=2).value) == 'Local Shunting Yard'):
        newtext.write('2')
    if(str(sheet1.cell(row=row_index, column=2).value) == 'Shunting Yard'):
        newtext.write('3')
    if(str(sheet1.cell(row=row_index, column=2).value) == 'Marshalling Yard'):
        newtext.write('4') 
    newtext.write('], \n\t')
    
newtext.write(']')

newtext.write('\n\n')

# --- ARCS ---

newtext.write('arc = [ \n\t')

for row_index in range(2, (sheet2.max_row+1)):
    #Forward
    newtext.write('[' + str(sheet2.cell(row=row_index, column=1).value[1:].lstrip('0'))  + ', ')   # Get arc starting node
    newtext.write(str(sheet2.cell(row=row_index, column=2).value[1:].lstrip('0')) + ', ')   # Get arc ending node
    newtext.write(str(sheet2.cell(row=row_index, column=3).value) + ', ') # Get its length/cost
    newtext.write(str(sheet2.cell(row=row_index, column=4).value)) # Get its capacity
    newtext.write('], \n\t')
    #Backward
    newtext.write('[' + str(sheet2.cell(row=row_index, column=2).value[1:].lstrip('0'))  + ', ')   # Get arc starting node
    newtext.write(str(sheet2.cell(row=row_index, column=1).value[1:].lstrip('0')) + ', ')   # Get arc ending node
    newtext.write(str(sheet2.cell(row=row_index, column=3).value) + ', ') # Get its length/cost
    newtext.write(str(sheet2.cell(row=row_index, column=4).value)) # Get its capacity
    newtext.write('], \n\t')

newtext.write(']')

newtext.write('\n\n')

# --- COMMODITIES ---

newtext.write('commodities = [ \n\t')

for row_index in range(2, (sheet3.max_row+1)):
    newtext.write('[' + str( int(sheet3.cell(row=row_index, column=1).value[1:].lstrip('0')) - 1 ) + ', ')   # Get commodity starting node
    newtext.write(str( int(sheet3.cell(row=row_index, column=2).value[1:].lstrip('0')) - 1 ) + ', ')   # Get commodity ending node
    newtext.write(str(sheet3.cell(row=row_index, column=3).value)) # Get commodity quantity
    newtext.write('], \n\t')
    
newtext.write(']')

newtext.write('\n\n')

# Do not forget to remove last comma of each array
newtext.close()
